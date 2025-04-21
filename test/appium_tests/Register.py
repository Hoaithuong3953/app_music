import unittest
import time
import requests
import logging
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import WebDriverException, NoSuchElementException, StaleElementReferenceException, TimeoutException, InvalidSessionIdException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Thiết lập logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

capabilities = dict(
    platformName='Android',
    automationName='uiautomator2',
    deviceName='emulator-5554',
    appPackage='com.example.app_music',
    appActivity='.MainActivity',
    language='en',
    locale='US',
    noReset=True,
    fullReset=False
)

appium_server_url = 'http://127.0.0.1:4723/wd/hub'

class TestRegister(unittest.TestCase):
    def setUp(self) -> None:
        print("🔄 Đang khởi động Appium driver...")
        logger.info("Đang khởi động Appium driver...")
        max_attempts = 3
        attempt = 1
        self.driver = None

        # Dọn dẹp phiên cũ
        try:
            session_id = capabilities.get("sessionId", "")
            if session_id:
                requests.delete(f'{appium_server_url}/session/{session_id}')
                print(f"✅ Đã xóa phiên cũ với sessionId: {session_id}")
                logger.info(f"Đã xóa phiên cũ với sessionId: {session_id}")
        except Exception as e:
            print(f"⚠️ Không thể xóa phiên cũ: {e}")
            logger.warning(f"Không thể xóa phiên cũ: {e}")

        # Thử lại khởi động driver
        while attempt <= max_attempts and not self.driver:
            try:
                options = UiAutomator2Options().load_capabilities(capabilities)
                self.driver = webdriver.Remote(appium_server_url, options=options)
                time.sleep(5)
                print("✅ Driver đã khởi động thành công.")
                logger.info("Driver đã khởi động thành công.")
                break
            except (WebDriverException, InvalidSessionIdException) as e:
                print(f"❌ Lỗi khi khởi động driver (lần {attempt}/{max_attempts}): {e}")
                logger.error(f"Lỗi khi khởi động driver (lần {attempt}/{max_attempts}): {e}")
                attempt += 1
                time.sleep(2)
                if attempt > max_attempts:
                    print("❌ Không thể khởi động driver sau nhiều lần thử.")
                    logger.error("Không thể khởi động driver sau nhiều lần thử.")
                    self.driver = None
                    raise

        # Khởi động lại ứng dụng
        try:
            self.driver.terminate_app("com.example.app_music")
            self.driver.activate_app("com.example.app_music")
            time.sleep(5)
            print("✅ Đã khởi động lại ứng dụng com.example.app_music.")
            logger.info("Đã khởi động lại ứng dụng com.example.app_music.")
        except Exception as e:
            print(f"⚠️ Lỗi khi khởi động lại ứng dụng: {e}")
            logger.error(f"Lỗi khi khởi động lại ứng dụng: {e}")
            self.driver = None
            raise

        # Chuẩn bị thư mục và file Excel
        self.excel_dir = "result"
        self.excel_file = os.path.join(self.excel_dir, "result_register.xlsx")
        self.init_excel()

    def init_excel(self):
        """Khởi tạo thư mục result và file Excel nếu chưa tồn tại"""
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
            print(f"✅ Đã tạo thư mục: {self.excel_dir}")
            logger.info(f"Đã tạo thư mục: {self.excel_dir}")

        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Register Results"
            ws.append(["Test Case", "First Name", "Last Name", "Mobile", "Email", "Password", "Result", "Status", "Timestamp"])
            wb.save(self.excel_file)
        print(f"✅ File Excel: {self.excel_file}")
        logger.info(f"File Excel: {self.excel_file}")

    def save_to_excel(self, test_case, first_name, last_name, mobile, email, password, result, status):
        """Lưu kết quả vào file Excel"""
        try:
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
                ws = wb["Register Results"]
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Register Results"
                ws.append(["Test Case", "First Name", "Last Name", "Mobile", "Email", "Password", "Result", "Status", "Timestamp"])

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([test_case, first_name, last_name, mobile, email, password, result, status, timestamp])
            wb.save(self.excel_file)
            print(f"✅ Đã thêm kết quả vào Excel: {test_case}, {first_name}, {last_name}, {mobile}, {email}, {password}, {result}, {status}")
            logger.info(f"Đã thêm kết quả vào Excel: {test_case}, {first_name}, {last_name}, {mobile}, {email}, {password}, {result}, {status}")
        except Exception as e:
            print(f"⚠️ Lỗi khi lưu vào Excel: {e}")
            logger.error(f"Lỗi khi lưu vào Excel: {e}")

    def tearDown(self) -> None:
        if self.driver:
            try:
                self.driver.quit()
                print("✅ Đã đóng driver.")
                logger.info("Đã đóng driver.")
            except WebDriverException as e:
                print(f"⚠️ Không thể đóng driver: {e}")
                logger.warning(f"Không thể đóng driver: {e}")
        else:
            print("⚠️ Không có driver để đóng.")
            logger.warning("Không có driver để đóng.")

    def ensure_app_in_foreground(self):
        """Đảm bảo ứng dụng đang ở trạng thái mong đợi"""
        try:
            current_activity = self.driver.current_activity
            if current_activity != '.MainActivity':
                print(f"⚠️ Ứng dụng không ở trạng thái mong đợi. Current activity: {current_activity}")
                logger.warning(f"Ứng dụng không ở trạng thái mong đợi. Current activity: {current_activity}")
                self.driver.activate_app(capabilities['appPackage'])
                time.sleep(5)
                current_activity = self.driver.current_activity
                if current_activity != '.MainActivity':
                    self.fail(f"❌ Không thể đưa ứng dụng về trạng thái mong đợi. Current activity: {current_activity}")
                print("✅ Đã đưa ứng dụng về MainActivity.")
                logger.info("Đã đưa ứng dụng về MainActivity.")
            else:
                print("✅ Ứng dụng đang ở MainActivity.")
                logger.info("Ứng dụng đang ở MainActivity.")
        except Exception as e:
            self.fail(f"❌ Lỗi khi kiểm tra trạng thái ứng dụng: {e}")
            logger.error(f"Lỗi khi kiểm tra trạng thái ứng dụng: {e}")

    def navigate_to_register(self):
        """Điều hướng đến màn hình đăng ký"""
        try:
            signup_nav_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, '//android.widget.Button[@content-desc="Don\'t have an account? Sign up"]'))
            )
            signup_nav_button.click()
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((AppiumBy.XPATH, '//android.view.View[@content-desc="Create a New Account"]'))
            )
            print("✅ Đã vào màn hình đăng ký.")
            logger.info("Đã vào màn hình đăng ký.")
        except Exception as e:
            print(f"❌ Lỗi khi điều hướng đến màn hình đăng ký: {e}")
            logger.error(f"Lỗi khi điều hướng đến màn hình đăng ký: {e}")
            raise

    def check_register_screen(self):
        """Kiểm tra xem vẫn ở màn hình đăng ký"""
        try:
            WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((AppiumBy.XPATH, '//android.view.View[@content-desc="Create a New Account"]'))
            )
            return True
        except:
            print("⚠️ Không còn ở màn hình đăng ký!")
            logger.warning("Không còn ở màn hình đăng ký!")
            print("🔍 Page Source khi rời khỏi màn hình:")
            print(self.driver.page_source)
            logger.debug(f"Page Source khi rời khỏi màn hình: {self.driver.page_source}")
            return False

    def hide_keyboard_alternative(self):
        """Ẩn bàn phím bằng cách nhấn vào khu vực ngoài trường nhập liệu"""
        try:
            self.driver.tap([(50, 50)])
            time.sleep(1)
            print("✅ Đã ẩn bàn phím bằng cách nhấn ngoài trường nhập liệu.")
            logger.info("Đã ẩn bàn phím bằng cách nhấn ngoài trường nhập liệu.")
        except Exception as e:
            print(f"⚠️ Lỗi khi ẩn bàn phím: {e}")
            logger.warning(f"Lỗi khi ẩn bàn phím: {e}")

    def input_field_with_retry(self, field_xpath, value, field_name, retries=5, wait_seconds=30, is_password=False):
        """Nhập liệu và kiểm tra với retry"""
        max_nav_attempts = 3
        nav_attempts = 0
        while nav_attempts < max_nav_attempts:
            if not self.check_register_screen():
                print("🔄 Màn hình đăng ký không còn, thử quay lại...")
                logger.info("Màn hình đăng ký không còn, thử quay lại...")
                self.restart_app()
                self.navigate_to_register()
                nav_attempts += 1
                continue
            for attempt in range(retries):
                try:
                    print(f"🔄 Thử nhập {field_name} lần {attempt + 1}/{retries}...")
                    logger.info(f"Thử nhập {field_name} lần {attempt + 1}/{retries}...")
                    field = WebDriverWait(self.driver, wait_seconds).until(
                        EC.element_to_be_clickable((AppiumBy.XPATH, field_xpath))
                    )
                    print(f"✅ Đã tìm thấy trường {field_name}.")
                    logger.info(f"Đã tìm thấy trường {field_name}.")
                    field.click()
                    time.sleep(1)
                    field.clear()
                    time.sleep(0.5)
                    if value:  # Chỉ nhập nếu giá trị không trống
                        field.send_keys(str(value))
                        time.sleep(1)
                        self.hide_keyboard_alternative()
                        field_text = field.get_attribute("text")
                        print(f"✅ Đã nhập {field_name}: {field_text}")
                        logger.info(f"Đã nhập {field_name}: {field_text}")
                        if not is_password:
                            self.assertEqual(field_text, str(value), f"{field_name} không được nhập đúng!")
                        else:
                            self.assertNotEqual(field_text, "", f"{field_name} không được nhập!")
                    else:
                        print(f"⏩ Bỏ qua nhập {field_name} vì giá trị trống.")
                        logger.info(f"Bỏ qua nhập {field_name} vì giá trị trống.")
                    if not self.check_register_screen():
                        print("🔍 Page Source sau khi nhập liệu:")
                        print(self.driver.page_source)
                        logger.debug(f"Page Source sau khi nhập liệu: {self.driver.page_source}")
                        self.fail(f"❌ Đã rời khỏi màn hình đăng ký sau khi nhập {field_name}!")
                    return
                except (StaleElementReferenceException, TimeoutException, WebDriverException) as e:
                    print(f"⚠️ Lỗi khi nhập {field_name} tại lần thử {attempt + 1}/{retries}: {e}")
                    logger.warning(f"Lỗi khi nhập {field_name} tại lần thử {attempt + 1}/{retries}: {e}")
                    print("🔍 Page Source khi gặp lỗi:")
                    print(self.driver.page_source)
                    logger.debug(f"Page Source khi gặp lỗi: {self.driver.page_source}")
                    time.sleep(2)
            self.fail(f"❌ Không thể nhập {field_name} sau {retries} lần thử.")
        self.fail(f"❌ Không thể quay lại màn hình đăng ký sau {max_nav_attempts} lần thử.")

    def scroll_to_element(self):
        """Cuộn để tìm nút Sign Up"""
        try:
            self.driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,
                                     'new UiScrollable(new UiSelector().scrollable(true)).scrollIntoView('
                                     'new UiSelector().description("Sign Up"))')
            print("✅ Đã cuộn đến nút Sign Up.")
            logger.info("Đã cuộn đến nút Sign Up.")
            return True
        except Exception as e:
            print(f"⚠️ Lỗi khi cuộn: {e}")
            logger.warning(f"Lỗi khi cuộn: {e}")
            return False

    def restart_app(self):
        """Khởi động lại ứng dụng nếu thoát"""
        try:
            self.driver.terminate_app('com.example.app_music')
            self.driver.activate_app('com.example.app_music')
            time.sleep(5)
            print("✅ Đã khởi động lại ứng dụng.")
            logger.info("Đã khởi động lại ứng dụng.")
        except Exception as e:
            print(f"❌ Lỗi khi khởi động lại ứng dụng: {e}")
            logger.error(f"Lỗi khi khởi động lại ứng dụng: {e}")
            raise

    def test_register_cases(self):
        if not self.driver:
            self.fail("❌ Driver chưa được khởi động.")
            logger.error("Driver chưa được khởi động.")

        # Kiểm tra xem ứng dụng đã khởi động đúng chưa
        self.ensure_app_in_foreground()

        # Kiểm tra xem có màn hình đăng nhập không
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((AppiumBy.CLASS_NAME, "android.widget.EditText"))
            )
            print("✅ Đã tìm thấy trường nhập liệu trên màn hình đăng nhập.")
            logger.info("Đã tìm thấy trường nhập liệu trên màn hình đăng nhập.")
        except Exception as e:
            print("🔍 In cấu trúc giao diện để debug:")
            print(self.driver.page_source)
            logger.error(f"Không tìm thấy màn hình đăng nhập: {e}")
            self.fail(f"❌ Không tìm thấy màn hình đăng nhập: {e}")

        # Lấy đường dẫn tuyệt đối của thư mục chứa file test_register.py
        current_dir = os.path.dirname(os.path.abspath(__file__))
        print(f"🔍 Đường dẫn hiện tại của file test_register.py: {current_dir}")
        logger.info(f"Đường dẫn hiện tại của file test_register.py: {current_dir}")

        # Điều hướng đến thư mục test_file (bên trong appium_tests)
        test_file_dir = os.path.join(current_dir, "test_file")
        test_file_dir = os.path.normpath(test_file_dir)
        print(f"🔍 Đường dẫn đến thư mục test_file: {test_file_dir}")
        logger.info(f"Đường dẫn đến thư mục test_file: {test_file_dir}")

        # Kiểm tra xem thư mục test_file có tồn tại không
        if not os.path.exists(test_file_dir):
            self.fail(f"❌ Thư mục test_file không tồn tại: {test_file_dir}")
            logger.error(f"Thư mục test_file không tồn tại: {test_file_dir}")

        # Xây dựng đường dẫn đến file test_register.xlsx
        test_file = os.path.join(test_file_dir, "test_register.xlsx")
        test_file = os.path.normpath(test_file)
        print(f"🔍 Đường dẫn file Excel: {test_file}")
        logger.info(f"Đường dẫn file Excel: {test_file}")

        # Kiểm tra xem file có tồn tại không
        if not os.path.exists(test_file):
            self.fail(f"❌ File Excel không tồn tại: {test_file}")
            logger.error(f"File Excel không tồn tại: {test_file}")

        # Đọc dữ liệu từ file test_register.xlsx
        try:
            wb = load_workbook(test_file)
            ws = wb.active
            test_cases = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                test_cases.append({
                    "Test Case": row[0],
                    "FirstName": row[1] if row[1] else "",
                    "LastName": row[2] if row[2] else "",
                    "Mobile": row[3] if row[3] else "",
                    "Email": row[4] if row[4] else "",
                    "Password": row[5] if row[5] else "",
                    "Expected Result": row[6]
                })
            print(f"✅ Đã đọc {len(test_cases)} test cases từ {test_file}")
            logger.info(f"Đã đọc {len(test_cases)} test cases từ {test_file}")
            print("🔍 Dữ liệu test cases:")
            logger.info("Dữ liệu test cases:")
            for test in test_cases:
                print(f"  - {test}")
                logger.info(f"  - {test}")
        except Exception as e:
            self.fail(f"❌ Lỗi khi đọc file Excel: {e}")
            logger.error(f"Lỗi khi đọc file Excel: {e}")

        # Chạy từng test case
        for test in test_cases:
            test_case = test["Test Case"]
            first_name = test["FirstName"]
            last_name = test["LastName"]
            mobile = test["Mobile"]
            email = test["Email"]
            password = test["Password"]
            expected_result = test["Expected Result"]
            print(f"🔄 Đang chạy test case: {test_case}")
            logger.info(f"Đang chạy test case: {test_case}")

            try:
                # Điều hướng đến màn hình đăng ký
                self.navigate_to_register()

                # Nhập thông tin (bỏ qua nếu trống)
                if first_name:
                    self.input_field_with_retry('//android.widget.EditText[@index="3"]', first_name, "First Name")
                    self.ensure_app_in_foreground()
                    if not self.check_register_screen():
                        self.fail("❌ Ứng dụng đã thoát khỏi màn hình đăng ký sau khi nhập First Name!")
                if last_name:
                    self.input_field_with_retry('//android.widget.EditText[@index="4"]', last_name, "Last Name")
                    self.ensure_app_in_foreground()
                    if not self.check_register_screen():
                        self.fail("❌ Ứng dụng đã thoát khỏi màn hình đăng ký sau khi nhập Last Name!")
                if mobile:
                    self.input_field_with_retry('//android.widget.EditText[@index="6"]', mobile, "Mobile")
                    self.ensure_app_in_foreground()
                    if not self.check_register_screen():
                        self.fail("❌ Ứng dụng đã thoát khỏi màn hình đăng ký sau khi nhập Mobile!")
                if email:
                    self.input_field_with_retry('//android.widget.EditText[@index="5"]', email, "Email")
                    self.ensure_app_in_foreground()
                    if not self.check_register_screen():
                        self.fail("❌ Ứng dụng đã thoát khỏi màn hình đăng ký sau khi nhập Email!")
                if password:
                    self.input_field_with_retry('//android.widget.EditText[@index="7"]', password, "Password", is_password=True)
                    self.ensure_app_in_foreground()
                    if not self.check_register_screen():
                        self.fail("❌ Ứng dụng đã thoát khỏi màn hình đăng ký sau khi nhập Password!")

                # Cuộn đến nút Sign Up
                if not self.scroll_to_element():
                    raise NoSuchElementException("Không thể cuộn để tìm nút Sign Up.")

                # Tìm và nhấn nút Sign Up
                signup_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, '//android.widget.Button[@content-desc="Sign Up"]'))
                )
                signup_button.click()
                print(f"✅ Đã nhấn nút Sign Up cho {test_case}.")
                logger.info(f"Đã nhấn nút Sign Up cho {test_case}.")
                time.sleep(5)

                # Kiểm tra kết quả
                try:
                    error_message = WebDriverWait(self.driver, 5).until(
                        EC.visibility_of_element_located(
                            (AppiumBy.XPATH, "//*[contains(@content-desc, 'không') or contains(@content-desc, 'đã tồn tại') or contains(@content-desc, 'hợp lệ') or contains(@content-desc, 'quá ngắn')]")
                        )
                    )
                    actual_result = error_message.get_attribute('content-desc')
                    status = "PASSED" if actual_result == expected_result else "FAILED"
                    print(f"🔍 Kết quả: {actual_result}, Trạng thái: {status}")
                    logger.info(f"Kết quả: {actual_result}, Trạng thái: {status}")
                except TimeoutException:
                    actual_result = "Thành Công"
                    status = "PASSED" if expected_result == "Thành Công" else "FAILED"
                    print(f"🔍 Kết quả: {actual_result}, Trạng thái: {status}")
                    logger.info(f"Kết quả: {actual_result}, Trạng thái: {status}")

                # Lưu kết quả vào Excel
                self.save_to_excel(
                    test_case=test_case,
                    first_name=first_name,
                    last_name=last_name,
                    mobile=mobile,
                    email=email,
                    password=password,
                    result=actual_result,
                    status=status
                )

                # Khởi động lại ứng dụng để đảm bảo trạng thái sạch cho test case tiếp theo
                self.restart_app()

            except NoSuchElementException as e:
                print("🔍 In cấu trúc giao diện để debug:")
                print(self.driver.page_source)
                logger.error(f"Không tìm thấy phần tử cần thiết cho {test_case}: {e}")
                self.save_to_excel(
                    test_case=test_case,
                    first_name=first_name,
                    last_name=last_name,
                    mobile=mobile,
                    email=email,
                    password=password,
                    result=f"Lỗi: {str(e)}",
                    status="FAILED"
                )
                self.fail(f"❌ Không tìm thấy phần tử cần thiết cho {test_case}: {e}")
            except Exception as e:
                print("🔍 In cấu trúc giao diện để debug:")
                print(self.driver.page_source)
                logger.error(f"Lỗi không xác định cho {test_case}: {e}")
                self.save_to_excel(
                    test_case=test_case,
                    first_name=first_name,
                    last_name=last_name,
                    mobile=mobile,
                    email=email,
                    password=password,
                    result=f"Lỗi: {str(e)}",
                    status="FAILED"
                )
                self.fail(f"❌ Lỗi không xác định cho {test_case}: {e}")

if __name__ == '__main__':
    unittest.main(verbosity=2)
import unittest
import time
import os
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import WebDriverException, NoSuchElementException, TimeoutException, NoSuchDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime

capabilities = dict(
    platformName='Android',
    automationName='uiautomator2',
    deviceName='emulator-5554',
    appPackage='com.example.app_music',
    appActivity='.MainActivity',
    language='en',
    locale='US',
    noReset=True
)

appium_server_url = 'http://127.0.0.1:4723/wd/hub'

class TestLogOut(unittest.TestCase):
    def setUp(self) -> None:
        print("🔄 Đang khởi động Appium driver...")
        try:
            options = UiAutomator2Options().load_capabilities(capabilities)
            self.driver = webdriver.Remote(appium_server_url, options=options)
            time.sleep(5)
            print("✅ Driver đã khởi động thành công.")
        except (WebDriverException, NoSuchDriverException) as e:
            print(f"❌ Lỗi khi khởi động driver: {e}")
            self.driver = None
            raise

        # Chuẩn bị thư mục và file Excel
        self.excel_dir = r"D:\Nam3\hocky2\kiemthu_giuaky\app_music\test\appium_tests\result"
        self.excel_file = os.path.join(self.excel_dir, "test_logout.xlsx")
        self.init_excel()

    def init_excel(self):
        """Khởi tạo thư mục result và file Excel nếu chưa tồn tại"""
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
            print(f"✅ Đã tạo thư mục: {self.excel_dir}")

        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Logout Results"
            ws.append(["Test Case", "Result", "Status", "Timestamp"])
            wb.save(self.excel_file)
        print(f"✅ File Excel: {self.excel_file}")

    def save_to_excel(self, test_case, result, status):
        """Lưu kết quả vào file Excel"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb["Logout Results"]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([test_case, result, status, timestamp])
            wb.save(self.excel_file)
            print(f"✅ Đã lưu kết quả vào Excel: {test_case}, {result}, {status}")
        except Exception as e:
            print(f"⚠️ Lỗi khi lưu vào Excel: {e}")

    def tearDown(self) -> None:
        if self.driver:
            try:
                self.driver.quit()
                print("✅ Đã đóng driver.")
            except (WebDriverException, NoSuchDriverException) as e:
                print(f"⚠️ Không thể đóng driver: {e}")
        else:
            print("⚠️ Không có driver để đóng.")

    def restart_app(self):
        """Khởi động lại ứng dụng nếu cần"""
        try:
            self.driver.terminate_app('com.example.app_music')
            self.driver.activate_app('com.example.app_music')
            time.sleep(5)
            print("✅ Đã khởi động lại ứng dụng.")
        except Exception as e:
            print(f"❌ Lỗi khi khởi động lại ứng dụng: {e}")
            raise

    def test_logout(self):
        """Kiểm tra chức năng đăng xuất từ ProfileScreen"""
        if not self.driver:
            self.fail("❌ Driver chưa được khởi động.")

        try:
            # Kiểm tra xem đang ở HomeScreen (MainActivity)
            print("🔍 Đang kiểm tra HomeScreen...")
            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Hi,')]")),
                message="Không tìm thấy tiêu đề HomeScreen"
            )
            print("✅ Đã vào HomeScreen.")

            # Tìm và nhấn nút Profile trong BottomNavigationBar
            print("🔍 Đang tìm nút Profile trong BottomNavigationBar...")
            profile_button = WebDriverWait(self.driver, 15).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//*[contains(@content-desc, 'Profile')]")),
                message="Không tìm thấy nút Profile"
            )
            profile_button.click()
            print("✅ Đã nhấn nút Profile.")

            # Kiểm tra đã vào ProfileScreen
            print("🔍 Đang kiểm tra ProfileScreen...")
            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Edit Profile')]")),
                message="Không tìm thấy tiêu đề ProfileScreen"
            )
            print("✅ Đã vào ProfileScreen.")

            # Lưu page_source để debug
            page_source = self.driver.page_source
            debug_file = os.path.join(self.excel_dir, "profile_screen_page_source.xml")
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(page_source)
            print(f"🔍 Đã lưu cấu trúc giao diện vào: {debug_file}")

            # Tìm và nhấn nút Log Out
            print("🔍 Đang tìm nút Log Out...")
            logout_button = WebDriverWait(self.driver, 15).until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//*[contains(@content-desc, 'Log Out')]")),
                message="Không tìm thấy nút Log Out"
            )
            logout_button.click()
            print("✅ Đã nhấn nút Log Out.")
            time.sleep(5)  # Chờ điều hướng

            # Kiểm tra đã chuyển về LoginScreen
            print("🔍 Đang kiểm tra LoginScreen...")
            login_screen_element = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Login')]")),
                message="Không chuyển được về LoginScreen"
            )
            self.assertTrue(login_screen_element.is_displayed(), "Không chuyển được về LoginScreen!")
            result = "Đăng xuất thành công và chuyển về LoginScreen"
            status = "PASSED"
            print("✅ Đăng xuất thành công và chuyển về LoginScreen!")

            # Lưu kết quả vào Excel
            self.save_to_excel(
                test_case="Log Out",
                result=result,
                status=status
            )

        except (NoSuchElementException, TimeoutException) as e:
            print("🔍 In cấu trúc giao diện để debug:")
            page_source = self.driver.page_source
            debug_file = os.path.join(self.excel_dir, "error_page_source.xml")
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(page_source)
            print(f"🔍 Đã lưu cấu trúc giao diện lỗi vào: {debug_file}")
            self.save_to_excel(
                test_case="Log Out",
                result=f"Lỗi: {str(e)}",
                status="FAILED"
            )
            self.fail(f"❌ Lỗi: {e}")
        except NoSuchDriverException as e:
            print("🔍 Session không hợp lệ, thử khởi động lại ứng dụng...")
            self.restart_app()
            self.save_to_excel(
                test_case="Log Out",
                result=f"Session không hợp lệ: {str(e)}",
                status="FAILED"
            )
            self.fail(f"❌ Session không hợp lệ: {e}")
        except Exception as e:
            print("🔍 In cấu trúc giao diện để debug:")
            page_source = self.driver.page_source
            debug_file = os.path.join(self.excel_dir, "error_page_source.xml")
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(page_source)
            print(f"🔍 Đã lưu cấu trúc giao diện lỗi vào: {debug_file}")
            self.save_to_excel(
                test_case="Log Out",
                result=f"Lỗi không xác định: {str(e)}",
                status="FAILED"
            )
            self.fail(f"❌ Lỗi không xác định: {e}")

if __name__ == '__main__':
    unittest.main(verbosity=2)
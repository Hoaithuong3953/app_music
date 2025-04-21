import unittest
import time
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import WebDriverException, NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

capabilities = dict(
    platformName='Android',
    automationName='uiautomator2',
    deviceName='emulator-5554',
    appPackage='com.example.app_music',
    appActivity='.MainActivity',
    language='en',
    locale='US',
    noReset=True,  # Không reset ứng dụng giữa các test
    fullReset=False  # Không reset hoàn toàn trạng thái ứng dụng
)

appium_server_url = 'http://127.0.0.1:4723/wd/hub'

class TestLoginFunction(unittest.TestCase):
    def setUp(self) -> None:
        print("🔄 Đang khởi động Appium driver...")
        try:
            options = UiAutomator2Options().load_capabilities(capabilities)
            self.driver = webdriver.Remote(appium_server_url, options=options)
            time.sleep(5)  # Chờ ứng dụng khởi động
            print("✅ Driver đã khởi động thành công.")
        except WebDriverException as e:
            print(f"❌ Lỗi khi khởi động driver: {e}")
            self.driver = None
            raise

        # Chuẩn bị thư mục và file Excel
        self.excel_dir = "result"
        self.excel_file = os.path.join(self.excel_dir, "result_login.xlsx")
        self.init_excel()

    def init_excel(self):
        """Khởi tạo thư mục result và file Excel nếu chưa tồn tại"""
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
            print(f"✅ Đã tạo thư mục: {self.excel_dir}")

        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Login Results"
            ws.append(["Test Case", "Email", "Password", "Result", "Status", "Timestamp"])
            wb.save(self.excel_file)
        print(f"✅ File Excel: {self.excel_file}")

    def save_to_excel(self, test_case, email, password, result, status):
        """Lưu kết quả vào file Excel mà không ghi đè dữ liệu cũ"""
        try:
            # Mở file Excel hiện có, nếu không tồn tại thì tạo mới
            if os.path.exists(self.excel_file):
                wb = load_workbook(self.excel_file)
                ws = wb["Login Results"]
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Login Results"
                ws.append(["Test Case", "Email", "Password", "Result", "Status", "Timestamp"])

            # Thêm dữ liệu mới vào cuối sheet
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([test_case, email, password, result, status, timestamp])

            # Lưu file
            wb.save(self.excel_file)
            print(f"✅ Đã thêm kết quả vào Excel: {test_case}, {email}, {password}, {result}, {status}")
        except Exception as e:
            print(f"⚠️ Lỗi khi lưu vào Excel: {e}")

    def tearDown(self) -> None:
        if self.driver:
            try:
                self.driver.quit()
                print("✅ Đã đóng driver.")
            except WebDriverException as e:
                print(f"⚠️ Không thể đóng driver: {e}")
        else:
            print("⚠️ Không có driver để đóng.")

    def scroll_to_element(self):
        """Cuộn để tìm nút Login dựa trên content-desc, trả về True nếu tìm thấy"""
        try:
            self.driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,
                                     'new UiScrollable(new UiSelector().scrollable(true)).scrollIntoView('
                                     'new UiSelector().description("Login"))')
            print("✅ Đã cuộn và tìm thấy nút Login.")
            return True
        except Exception as e:
            print(f"⚠️ Không tìm thấy nút Login khi cuộn: {e}")
            return False

    def hide_keyboard_alternative(self):
        """Ẩn bàn phím bằng cách nhấn vào khu vực ngoài trường nhập liệu"""
        try:
            # Nhấn vào một vị trí ngoài trường nhập liệu (góc trên bên trái màn hình)
            self.driver.tap([(50, 50)])
            time.sleep(1)
            print("✅ Đã ẩn bàn phím bằng cách nhấn ngoài trường nhập liệu.")
        except Exception as e:
            print(f"⚠️ Lỗi khi ẩn bàn phím: {e}")

    def ensure_app_in_foreground(self):
        """Đảm bảo ứng dụng đang ở trạng thái mong đợi"""
        try:
            current_activity = self.driver.current_activity
            if current_activity != '.MainActivity':
                print(f"⚠️ Ứng dụng không ở trạng thái mong đợi. Current activity: {current_activity}")
                # Khởi động lại ứng dụng
                self.driver.activate_app(capabilities['appPackage'])
                time.sleep(5)
                # Kiểm tra lại
                current_activity = self.driver.current_activity
                if current_activity != '.MainActivity':
                    self.fail(f"❌ Không thể đưa ứng dụng về trạng thái mong đợi. Current activity: {current_activity}")
                print("✅ Đã đưa ứng dụng về MainActivity.")
            else:
                print("✅ Ứng dụng đang ở MainActivity.")
        except Exception as e:
            self.fail(f"❌ Lỗi khi kiểm tra trạng thái ứng dụng: {e}")

    def test_login_cases(self):
        if not self.driver:
            self.fail("❌ Driver chưa được khởi động.")

        # Kiểm tra xem ứng dụng đã khởi động đúng chưa
        self.ensure_app_in_foreground()

        # Kiểm tra xem có màn hình đăng nhập không
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((AppiumBy.CLASS_NAME, "android.widget.EditText"))
            )
            print("✅ Đã tìm thấy trường nhập liệu, sẵn sàng để nhập thông tin đăng nhập.")
        except Exception as e:
            print("🔍 In cấu trúc giao diện để debug:")
            print(self.driver.page_source)
            self.fail(f"❌ Không tìm thấy màn hình đăng nhập: {e}")

        # Lấy đường dẫn tuyệt đối của thư mục chứa file Login.py
        current_dir = os.path.dirname(os.path.abspath(__file__))
        print(f"🔍 Đường dẫn hiện tại của file Login.py: {current_dir}")

        # Điều hướng đến thư mục test_file (bên trong appium_tests)
        test_file_dir = os.path.join(current_dir, "test_file")
        test_file_dir = os.path.normpath(test_file_dir)
        print(f"🔍 Đường dẫn đến thư mục test_file: {test_file_dir}")

        # Kiểm tra xem thư mục test_file có tồn tại không
        if not os.path.exists(test_file_dir):
            self.fail(f"❌ Thư mục test_file không tồn tại: {test_file_dir}")

        # Xây dựng đường dẫn đến file test_login.xlsx
        test_file = os.path.join(test_file_dir, "test_login.xlsx")
        test_file = os.path.normpath(test_file)
        print(f"🔍 Đường dẫn file Excel: {test_file}")

        # Kiểm tra xem file có tồn tại không
        if not os.path.exists(test_file):
            self.fail(f"❌ File Excel không tồn tại: {test_file}")

        # Đọc dữ liệu từ file test_login.xlsx
        try:
            wb = load_workbook(test_file)
            ws = wb.active
            test_cases = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                test_cases.append({
                    "Test Case": row[0],
                    "Email": row[1] if row[1] else "",
                    "Password": row[2] if row[2] else "",
                    "Expected Result": row[3]
                })
            print(f"✅ Đã đọc {len(test_cases)} test cases từ {test_file}")
            print("🔍 Dữ liệu test cases:")
            for test in test_cases:
                print(f"  - {test}")
        except Exception as e:
            self.fail(f"❌ Lỗi khi đọc file Excel: {e}")

        # Chạy test case
        for test in test_cases:
            test_case = test["Test Case"]
            email = test["Email"]
            password = test["Password"]
            expected_result = test["Expected Result"]
            print(f"🔄 Đang chạy test case: {test_case}")

            try:
                # Tìm và nhập email
                email_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((AppiumBy.CLASS_NAME, "android.widget.EditText"))
                )[0]  # Index 0 cho email
                email_field.click()
                email_field.clear()
                if email:
                    email_field.send_keys(email)
                time.sleep(1)
                email_text = email_field.get_attribute("text")
                print(f"✅ Đã nhập email: {email_text}")
                self.assertEqual(email_text, email, f"Email không được nhập đúng cho {test_case}!")

                # Tìm và nhập mật khẩu
                password_field = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((AppiumBy.CLASS_NAME, "android.widget.EditText"))
                )[1]  # Index 1 cho password
                password_field.click()
                password_field.clear()
                if password:
                    password_field.send_keys(str(password))  # Chuyển password thành chuỗi
                time.sleep(1)
                password_text = password_field.get_attribute("text")
                print(f"✅ Đã nhập mật khẩu: {password_text}")
                self.assertEqual(len(password_text), len(str(password)), f"Độ dài mật khẩu không khớp cho {test_case}!")

                # Ẩn bàn phím bằng cách nhấn ngoài trường nhập liệu
                self.hide_keyboard_alternative()

                # Kiểm tra lại trạng thái ứng dụng sau khi nhập liệu
                self.ensure_app_in_foreground()

                # Cuộn đến nút Login
                if not self.scroll_to_element():
                    raise NoSuchElementException("Không thể cuộn để tìm nút Login.")

                # Tìm và nhấn nút Login
                login_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@content-desc='Login']"))
                )
                login_button.click()
                print(f"✅ Đã nhấn nút Login cho {test_case}.")
                time.sleep(3)

                # Kiểm tra kết quả
                try:
                    error_message = WebDriverWait(self.driver, 5).until(
                        EC.visibility_of_element_located(
                            (AppiumBy.XPATH, "//*[@content-desc[contains(., 'không') or contains(., 'Sai') or contains(., 'thất bại')]]")
                        )
                    )
                    actual_result = error_message.get_attribute('content-desc')
                    status = "PASSED" if actual_result == expected_result else "FAILED"
                    print(f"🔍 Kết quả: {actual_result}, Trạng thái: {status}")
                except TimeoutException:
                    actual_result = "Đăng nhập thành công (không có thông báo lỗi)"
                    status = "PASSED" if expected_result == "Đăng nhập thành công" else "FAILED"
                    print(f"🔍 Kết quả: {actual_result}, Trạng thái: {status}")

                # Lưu kết quả vào Excel
                self.save_to_excel(
                    test_case=test_case,
                    email=email,
                    password=str(password),  # Chuyển password thành chuỗi khi lưu
                    result=actual_result,
                    status=status
                )

                # Khởi động lại ứng dụng để đảm bảo trạng thái sạch cho test case tiếp theo
                self.driver.terminate_app(capabilities['appPackage'])
                self.driver.activate_app(capabilities['appPackage'])
                time.sleep(5)

            except NoSuchElementException as e:
                print("🔍 In cấu trúc giao diện để debug:")
                print(self.driver.page_source)
                self.save_to_excel(
                    test_case=test_case,
                    email=email,
                    password=str(password),  # Chuyển password thành chuỗi khi lưu
                    result=f"Lỗi: {str(e)}",
                    status="FAILED"
                )
                self.fail(f"❌ Không tìm thấy phần tử cần thiết cho {test_case}: {e}")
            except Exception as e:
                print("🔍 In cấu trúc giao diện để debug:")
                print(self.driver.page_source)
                self.save_to_excel(
                    test_case=test_case,
                    email=email,
                    password=str(password),  # Chuyển password thành chuỗi khi lưu
                    result=f"Lỗi: {str(e)}",
                    status="FAILED"
                )
                self.fail(f"❌ Lỗi không xác định cho {test_case}: {e}")

if __name__ == "__main__":
    unittest.main(verbosity=2)
import unittest
import time
import logging
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import WebDriverException, NoSuchElementException, TimeoutException, InvalidSessionIdException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import requests

# Thiết lập logging
logging.basicConfig(
    filename="test_search.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger()

capabilities = {
    "platformName": "Android",
    "appium:automationName": "uiautomator2",
    "appium:deviceName": "emulator-5554",
    "appium:appPackage": "com.example.app_music",
    "appium:appActivity": ".MainActivity",
    "appium:language": "en",
    "appium:locale": "US",
    "appium:newCommandTimeout": 300,
    "appium:noReset": False,
    "appium:adbExecTimeout": 60000,
    "appium:appWaitActivity": "*"
}

appium_server_url = 'http://127.0.0.1:4723/wd/hub'

class SearchMusicTest(unittest.TestCase):
    def setUp(self):
        print("🔄 Đang khởi động Appium driver...")
        logger.info("Đang khởi động Appium driver...")
        max_attempts = 3
        attempt = 1
        self.driver = None

        # Dọn dẹp phiên cũ
        try:
            requests.delete(f'{appium_server_url}/session/{capabilities.get("sessionId", "")}')
        except Exception as e:
            print(f"⚠️ Không thể xóa phiên cũ: {e}")
            logger.warning(f"Không thể xóa phiên cũ: {e}")

        # Thử lại khởi động driver
        while attempt <= max_attempts and not self.driver:
            try:
                options = UiAutomator2Options().load_capabilities(capabilities)
                self.driver = webdriver.Remote(appium_server_url, options=options)
                time.sleep(5)
                print("✅ Driver đã khởi động thành công.")
                logger.info("Driver đã khởi động thành công.")
                break
            except (WebDriverException, InvalidSessionIdException) as e:
                print(f"❌ Lỗi khi khởi động driver (lần {attempt}/{max_attempts}): {e}")
                logger.error(f"Lỗi khi khởi động driver (lần {attempt}/{max_attempts}): {e}")
                attempt += 1
                time.sleep(2)
                if attempt > max_attempts:
                    print("❌ Không thể khởi động driver sau nhiều lần thử.")
                    logger.error("Không thể khởi động driver sau nhiều lần thử.")
                    self.driver = None
                    raise

        self.wait = WebDriverWait(self.driver, 40)

        # Khởi động lại ứng dụng
        self.driver.terminate_app("com.example.app_music")
        self.driver.activate_app("com.example.app_music")
        time.sleep(5)

        # Chuẩn bị thư mục và file Excel
        self.excel_dir = "result"
        self.excel_file = os.path.join(self.excel_dir, "result.xlsx")
        self.init_excel()

    def init_excel(self):
        """Khởi tạo thư mục result và file Excel nếu chưa tồn tại"""
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
            print(f"✅ Đã tạo thư mục: {self.excel_dir}")
            logger.info(f"Đã tạo thư mục: {self.excel_dir}")

        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Search Results"
            ws.append(["Test Case", "Search Query", "Result", "Status", "Timestamp"])
            wb.save(self.excel_file)
        print(f"✅ File Excel: {self.excel_file}")
        logger.info(f"File Excel: {self.excel_file}")

    def save_to_excel(self, test_case, query, result, status):
        """Lưu kết quả vào file Excel"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb["Search Results"]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([test_case, query, result, status, timestamp])
            wb.save(self.excel_file)
            print(f"✅ Đã lưu kết quả vào Excel: {test_case}, {query}, {result}, {status}")
            logger.info(f"Đã lưu kết quả vào Excel: {test_case}, {query}, {result}, {status}")
        except Exception as e:
            print(f"⚠️ Lỗi khi lưu vào Excel: {e}")
            logger.error(f"Lỗi khi lưu vào Excel: {e}")

    def tearDown(self):
        if self.driver:
            try:
                self.driver.quit()
                print("✅ Đã đóng driver.")
                logger.info("Đã đóng driver.")
            except (WebDriverException, InvalidSessionIdException) as e:
                print(f"⚠️ Không thể đóng driver: {e}")
                logger.error(f"Không thể đóng driver: {e}")
        else:
            print("⚠️ Không có driver để đóng.")
            logger.warning("Không có driver để đóng.")

    def restart_app(self):
        """Khởi động lại ứng dụng nếu cần"""
        try:
            self.driver.terminate_app('com.example.app_music')
            self.driver.activate_app('com.example.app_music')
            time.sleep(5)
            print("✅ Đã khởi động lại ứng dụng.")
            logger.info("Đã khởi động lại ứng dụng.")
        except Exception as e:
            print(f"❌ Lỗi khi khởi động lại ứng dụng: {e}")
            logger.error(f"Lỗi khi khởi động lại ứng dụng: {e}")
            raise

    def ensure_app_active(self):
        """Đảm bảo ứng dụng đang ở trạng thái hoạt động"""
        try:
            current_activity = self.driver.current_activity
            if current_activity != ".MainActivity":
                print(f"⚠️ Ứng dụng không ở trạng thái mong muốn: {current_activity}. Khởi động lại...")
                logger.warning(f"Ứng dụng không ở trạng thái mong muốn: {current_activity}. Khởi động lại...")
                self.restart_app()
                self.wait.until(
                    EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Hi,')]")),
                    message="Không thể quay lại màn hình chính"
                )
                print("✅ Đã khởi động lại ứng dụng.")
                logger.info("Đã khởi động lại ứng dụng.")
        except Exception as e:
            print(f"❌ Lỗi khi kiểm tra trạng thái ứng dụng: {e}")
            logger.error(f"Lỗi khi kiểm tra trạng thái ứng dụng: {e}")
            raise

    def search_and_check(self, query, test_case):
        """Tìm kiếm và kiểm tra kết quả, lưu vào Excel"""
        try:
            print(f"🔍 Đang kiểm tra trạng thái ứng dụng trước khi tìm kiếm '{query}'...")
            logger.info(f"Đang kiểm tra trạng thái ứng dụng trước khi tìm kiếm '{query}'...")
            self.ensure_app_active()
            current_activity = self.driver.current_activity
            print(f"✅ Trạng thái ứng dụng: {current_activity}")
            logger.info(f"Trạng thái ứng dụng: {current_activity}")

            print(f"🔍 Đang tìm thanh tìm kiếm cho '{query}'...")
            logger.info(f"Đang tìm thanh tìm kiếm cho '{query}'...")
            search_field = self.wait.until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.EditText[1]")),
                message="Không tìm thấy thanh tìm kiếm"
            )
            print(f"🔍 Thuộc tính thanh tìm kiếm: enabled={search_field.is_enabled()}, displayed={search_field.is_displayed()}")
            logger.info(f"Thuộc tính thanh tìm kiếm: enabled={search_field.is_enabled()}, displayed={search_field.is_displayed()}")

            if not search_field.is_enabled() or not search_field.is_displayed():
                raise Exception("Thanh tìm kiếm không ở trạng thái có thể tương tác.")

            print("🔍 Giao diện trước khi nhấn thanh tìm kiếm:")
            logger.info("Giao diện trước khi nhấn thanh tìm kiếm:")
            print(self.driver.page_source)
            logger.info(self.driver.page_source)

            search_field.click()
            time.sleep(1)
            print("✅ Đã nhấn vào thanh tìm kiếm.")
            logger.info("Đã nhấn vào thanh tìm kiếm.")

            search_field.clear()
            print(f"✅ Đã xóa nội dung cũ trong thanh tìm kiếm trước khi nhập '{query}'.")
            logger.info(f"Đã xóa nội dung cũ trong thanh tìm kiếm trước khi nhập '{query}'.")

            print(f"🔍 Đang nhập '{query}' vào thanh tìm kiếm...")
            logger.info(f"Đang nhập '{query}' vào thanh tìm kiếm...")
            try:
                self.driver.execute_script("mobile: shell", {"command": f"input text '{query}'"})
                print(f"✅ Đã nhập '{query}' bằng ADB.")
                logger.info(f"Đã nhập '{query}' bằng ADB.")
            except Exception as e:
                print(f"⚠️ Lỗi khi nhập bằng ADB, thử nhập từng ký tự: {e}")
                logger.warning(f"Lỗi khi nhập bằng ADB, thử nhập từng ký tự: {e}")
                for char in query:
                    search_field.send_keys(char)
                    time.sleep(0.3)
                    current_activity = self.driver.current_activity
                    if not current_activity:
                        raise Exception(f"Ứng dụng đã thoát khi nhập ký tự '{char}'.")
                    print(f"✅ Đã nhập ký tự: {char}")
                    logger.info(f"Đã nhập ký tự: {char}")
                print(f"✅ Đã nhập '{query}' bằng từng ký tự.")
                logger.info(f"Đã nhập '{query}' bằng từng ký tự.")

            current_activity = self.driver.current_activity
            if not current_activity:
                raise Exception("Ứng dụng đã thoát sau khi nhập thông tin.")
            print(f"✅ Trạng thái ứng dụng sau khi nhập: {current_activity}")
            logger.info(f"Trạng thái ứng dụng sau khi nhập: {current_activity}")

            print("🔍 Giao diện sau khi nhập thông tin:")
            logger.info("Giao diện sau khi nhập thông tin:")
            print(self.driver.page_source)
            logger.info(self.driver.page_source)

            time.sleep(20)  # Chờ lâu hơn để kết quả tải
            print(f"✅ Đã chờ 20 giây sau khi nhập '{query}'. Trạng thái: {current_activity}")
            logger.info(f"Đã chờ 20 giây sau khi nhập '{query}'. Trạng thái: {current_activity}")

            # Debug: In tất cả TextView để kiểm tra kết quả
            print("🔍 In tất cả TextView để debug:")
            logger.info("In tất cả TextView để debug:")
            text_views = self.driver.find_elements(AppiumBy.XPATH, "//android.widget.TextView")
            for tv in text_views:
                text = tv.get_attribute("text")
                print(f"TextView: {text}")
                logger.info(f"TextView: {text}")

            # Kiểm tra kết quả tìm kiếm
            print("🔍 Kiểm tra kết quả tìm kiếm...")
            logger.info("Kiểm tra kết quả tìm kiếm...")
            try:
                # Tìm tất cả SongCard dựa trên nút play (trailing)
                result_elements = self.driver.find_elements(
                    AppiumBy.XPATH,
                    "//android.view.ViewGroup[.//android.widget.ImageButton[@content-desc='play_song_button']]"
                )
                valid_results = []
                for element in result_elements:
                    try:
                        # Tìm TextView con trong ViewGroup (title và artist)
                        text_views_in_card = element.find_elements(AppiumBy.XPATH, ".//android.widget.TextView")
                        if len(text_views_in_card) >= 2:  # Đảm bảo có ít nhất 2 TextView (title và artist)
                            title_text = text_views_in_card[0].get_attribute("text")
                            artist_text = text_views_in_card[1].get_attribute("text")
                            # Kiểm tra query có trong title hoặc artist
                            if query.lower() in title_text.lower() or query.lower() in artist_text.lower():
                                valid_results.append(element)
                                print(f"Found SongCard: Title={title_text}, Artist={artist_text}")
                                logger.info(f"Found SongCard: Title={title_text}, Artist={artist_text}")
                    except NoSuchElementException:
                        continue

                if valid_results:
                    result_count = len(valid_results)
                    result_text = f"Tìm thấy {result_count} bài hát chứa '{query}'"
                    status = "PASSED"
                    print(f"✅ {result_text}")
                    logger.info(f"{result_text}")
                    # Debug: In chi tiết các SongCard
                    for i, card in enumerate(valid_results, 1):
                        text_views_in_card = card.find_elements(AppiumBy.XPATH, ".//android.widget.TextView")
                        card_details = [tv.get_attribute("text") for tv in text_views_in_card]
                        print(f"SongCard {i}: {card_details}")
                        logger.info(f"SongCard {i}: {card_details}")
                else:
                    # Kiểm tra "No results found"
                    try:
                        no_result_element = self.wait.until(
                            EC.presence_of_element_located(
                                (AppiumBy.XPATH, "//android.widget.TextView[@content-desc='no_results_found']")
                            ),
                            message="Không tìm thấy trạng thái 'No results found'"
                        )
                        result_text = "No results found"
                        status = "PARTIAL (No results found)"
                        print(f"✅ {result_text}")
                        logger.info(f"{result_text}")
                    except TimeoutException:
                        # Thử tìm TextView với text trực tiếp nếu không có Semantics
                        try:
                            no_result_element = self.wait.until(
                                EC.presence_of_element_located(
                                    (AppiumBy.XPATH, "//android.widget.TextView[@text='No results found']")
                                ),
                                message="Không tìm thấy trạng thái 'No results found' (text)"
                            )
                            result_text = "No results found"
                            status = "PARTIAL (No results found)"
                            print(f"✅ {result_text}")
                            logger.info(f"{result_text}")
                        except TimeoutException:
                            result_text = f"Không tìm thấy kết quả hoặc trạng thái 'No results found' cho '{query}'"
                            status = "FAILED"
                            print(f"⚠️ {result_text}")
                            logger.info(f"{result_text}")

                # Xóa nội dung thanh tìm kiếm
                print(f"🔍 Xóa nội dung thanh tìm kiếm sau khi kiểm tra '{query}'...")
                logger.info(f"Xóa nội dung thanh tìm kiếm sau khi kiểm tra '{query}'...")
                current_activity = self.driver.current_activity
                if not current_activity:
                    raise Exception("Ứng dụng đã thoát trước khi xóa thanh tìm kiếm.")
                if not (search_field.is_enabled() and search_field.is_displayed()):
                    raise Exception("Thanh tìm kiếm không ở trạng thái có thể tương tác để xóa.")
                search_field.click()
                time.sleep(1)
                search_field.clear()
                print(f"✅ Đã xóa nội dung thanh tìm kiếm.")
                logger.info("Đã xóa nội dung thanh tìm kiếm.")

                self.save_to_excel(
                    test_case=test_case,
                    query=query,
                    result=result_text,
                    status=status
                )

            except Exception as e:
                print(f"⚠️ Lỗi khi kiểm tra kết quả cho '{query}': {e}")
                logger.error(f"Lỗi khi kiểm tra kết quả cho '{query}': {e}")
                result_text = f"Lỗi khi kiểm tra kết quả: {str(e)}"
                status = "FAILED"
                self.save_to_excel(
                    test_case=test_case,
                    query=query,
                    result=result_text,
                    status=status
                )

        except Exception as e:
            print(f"⚠️ Lỗi khi tìm kiếm '{query}' (có thể ứng dụng đã thoát): {e}")
            logger.error(f"Lỗi khi tìm kiếm '{query}' (có thể ứng dụng đã thoát): {e}")
            print("🔍 In cấu trúc giao diện để debug:")
            logger.info("In cấu trúc giao diện để debug:")
            try:
                print(self.driver.page_source)
                logger.info(self.driver.page_source)
            except:
                print("⚠️ Không thể lấy page_source, ứng dụng đã crash.")
                logger.error("Không thể lấy page_source, ứng dụng đã crash.")
            self.save_to_excel(
                test_case=test_case,
                query=query,
                result=f"Lỗi: {str(e)}",
                status="FAILED"
            )
            self.fail(f"❌ Lỗi không xác định: {e}")

    def test_sequential_search(self):
        """Tìm kiếm lần lượt Đánh đổi, Fly me to the moon, Obito và lưu kết quả vào Excel"""
        if not self.driver:
            self.fail("❌ Driver chưa được khởi động.")

        # Bước đăng nhập
        try:
            # Tìm trường email
            email_field = self.wait.until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.EditText[1]")),
                message="Không tìm thấy trường email"
            )
            email_field.click()
            email_field.clear()
            email_field.send_keys("khai@gmail.com")
            time.sleep(1)
            email_text = email_field.get_attribute("text")
            print(f"✅ Đã nhập email: {email_text}")
            logger.info(f"Đã nhập email: {email_text}")
            self.assertEqual(email_text, "khai@gmail.com", "Email không được nhập đúng!")

            # Tìm trường mật khẩu
            password_field = self.wait.until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.EditText[2]")),
                message="Không tìm thấy trường mật khẩu"
            )
            password_field.click()
            password_field.clear()
            password_field.send_keys("123456")
            time.sleep(1)
            password_text = password_field.get_attribute("text")
            print(f"✅ Đã nhập mật khẩu: {password_text}")
            logger.info(f"Đã nhập mật khẩu: {password_text}")
            self.assertEqual(len(password_text), len("123456"), "Độ dài mật khẩu không khớp!")

            # Nhấn nút Login
            login_button = self.wait.until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@content-desc='Login']")),
                message="Không tìm thấy nút Login"
            )
            login_button.click()
            time.sleep(10)  # Chờ màn hình chính tải
            print("✅ Đã nhấn nút Login")
            logger.info("Đã nhấn nút Login")

            # Chờ màn hình chính
            self.wait.until(
                EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Hi,')]")),
                message="Không tìm thấy tiêu đề HomeScreen"
            )
            print("✅ Đã vào màn hình chính.")
            logger.info("Đã vào màn hình chính.")
            print("🔍 Giao diện màn hình chính:")
            logger.info("Giao diện màn hình chính:")
            print(self.driver.page_source)
            logger.info(self.driver.page_source)

        except (NoSuchElementException, TimeoutException) as e:
            print("🔍 In cấu trúc giao diện để debug:")
            logger.info("In cấu trúc giao diện để debug:")
            print(self.driver.page_source)
            logger.info(self.driver.page_source)
            self.save_to_excel(
                test_case="Login",
                query="khai@gmail.com",
                result=f"Lỗi đăng nhập: {str(e)}",
                status="FAILED"
            )
            self.fail(f"❌ Lỗi khi đăng nhập: {e}")

        # Tìm kiếm lần lượt
        self.search_and_check("Đánh đổi", "Search Đánh đổi")
        self.search_and_check("Fly me to the moon", "Search Fly me to the moon")
        self.search_and_check("Obito", "Search Obito")

if __name__ == '__main__':
    unittest.main(verbosity=2)
import unittest
import time
import logging
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
import selenium.common.exceptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Thiết lập logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SearchMusicTest(unittest.TestCase):
    def setUp(self):
        """Khởi tạo Appium driver trước mỗi test case"""
        print("🔄 Đang khởi động Appium driver...")
        logger.info("Đang khởi động Appium driver...")
        max_attempts = 3
        attempt = 1
        self.driver = None

        while attempt <= max_attempts and not self.driver:
            try:
                desired_caps = {
                    "platformName": "Android",
                    "deviceName": "emulator-5554",
                    "appPackage": "com.example.app_music",
                    "appActivity": ".MainActivity",
                    "automationName": "UiAutomator2",
                    "newCommandTimeout": 300,
                    "adbExecTimeout": 60000,
                    "appWaitActivity": "*",
                    "noReset": True,
                    "fullReset": False,
                    "language": "en",
                    "locale": "US"
                }
                options = UiAutomator2Options().load_capabilities(desired_caps)
                self.driver = webdriver.Remote("http://127.0.0.1:4723/wd/hub", options=options)
                time.sleep(5)
                print("✅ Driver đã khởi động thành công.")
                logger.info("Driver đã khởi động thành công.")
                break
            except (selenium.common.exceptions.WebDriverException, selenium.common.exceptions.InvalidSessionIdException) as e:
                print(f"❌ Lỗi khi khởi động driver (lần {attempt}/{max_attempts}): {e}")
                logger.error(f"Lỗi khi khởi động driver (lần {attempt}/{max_attempts}): {e}")
                attempt += 1
                time.sleep(2)
                if attempt > max_attempts:
                    print("❌ Không thể khởi động driver sau nhiều lần thử.")
                    logger.error("Không thể khởi động driver sau nhiều lần thử.")
                    self.driver = None
                    raise

        self.wait = WebDriverWait(self.driver, 60)  # Tăng thời gian đợi lên 60 giây
        self.driver.terminate_app("com.example.app_music")
        self.driver.activate_app("com.example.app_music")
        time.sleep(5)
        print("✅ Đã khởi động lại ứng dụng com.example.app_music.")
        logger.info("Đã khởi động lại ứng dụng com.example.app_music.")

        # Chuẩn bị file Excel
        self.excel_file = r"D:\Nam3\hocky2\kiemthu_giuaky\app_music\test\appium_tests\result\play_pause_results.xlsx"
        self.init_excel()

    def init_excel(self):
        """Khởi tạo file Excel nếu chưa tồn tại"""
        os.makedirs(os.path.dirname(self.excel_file), exist_ok=True)
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Music Playback Results"
            ws.append(["Test Case", "Song Name", "Result", "Status", "Timestamp"])
            wb.save(self.excel_file)
        print(f"✅ File Excel: {self.excel_file}")
        logger.info(f"File Excel: {self.excel_file}")

    def save_to_excel(self, test_case, song_name, result, status):
        """Lưu kết quả vào file Excel"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb["Music Playback Results"]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([test_case, song_name, result, status, timestamp])
            wb.save(self.excel_file)
            print(f"✅ Đã lưu kết quả vào Excel: {test_case}, {song_name}, {result}, {status}")
            logger.info(f"Đã lưu kết quả vào Excel: {test_case}, {song_name}, {result}, {status}")
        except Exception as e:
            print(f"⚠️ Lỗi khi lưu vào Excel: {e}")
            logger.error(f"Lỗi khi lưu vào Excel: {e}")

    def ensure_app_foreground(self):
        """Đảm bảo ứng dụng ở foreground"""
        try:
            current_package = self.driver.current_package
            if current_package != 'com.example.app_music':
                print(f"⚠️ Ứng dụng không ở foreground: {current_package}. Kích hoạt lại...")
                logger.warning(f"Ứng dụng không ở foreground: {current_package}. Kích hoạt lại...")
                self.driver.activate_app('com.example.app_music')
                time.sleep(3)
                self.wait.until(
                    EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Hi,')]")),
                    message="Không trở lại HomeScreen"
                )
                print("✅ Đã trở lại ứng dụng.")
                logger.info("Đã trở lại ứng dụng.")
        except Exception as e:
            print(f"⚠️ Lỗi khi kiểm tra foreground: {e}")
            logger.error(f"Lỗi khi kiểm tra foreground: {e}")

    def login(self):
        """Đăng nhập vào ứng dụng"""
        try:
            print("🔍 Bắt đầu đăng nhập...")
            logger.info("Bắt đầu đăng nhập...")
            email_field = self.wait.until(
                EC.presence_of_all_elements_located((AppiumBy.CLASS_NAME, "android.widget.EditText"))
            )[0]
            email_field.click()
            email_field.clear()
            email_field.send_keys("khai@gmail.com")
            time.sleep(1)
            print("✅ Đã nhập email.")
            logger.info("Đã nhập email.")

            password_field = self.wait.until(
                EC.presence_of_all_elements_located((AppiumBy.CLASS_NAME, "android.widget.EditText"))
            )[1]
            password_field.click()
            password_field.clear()
            password_field.send_keys("123456")
            time.sleep(1)
            print("✅ Đã nhập mật khẩu.")
            logger.info("Đã nhập mật khẩu.")

            self.driver.find_element(
                AppiumBy.ANDROID_UIAUTOMATOR,
                'new UiScrollable(new UiSelector().scrollable(true)).scrollIntoView('
                'new UiSelector().description("Login"))'
            )
            login_button = self.wait.until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@content-desc='Login']"))
            )
            login_button.click()
            print("✅ Đã nhấn nút Login.")
            logger.info("Đã nhấn nút Login.")

            self.wait.until(
                EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Hi,')]")),
                message="Không tìm thấy tiêu đề HomeScreen"
            )
            time.sleep(2)
            print("✅ Đã vào màn hình chính.")
            logger.info("Đã vào màn hình chính.")
        except Exception as e:
            print("🔍 In cấu trúc giao diện để debug:")
            print(self.driver.page_source)
            logger.error(f"Lỗi đăng nhập: {e}")
            self.fail(f"❌ Lỗi đăng nhập: {e}")

    def scroll_to_recommended_songs(self):
        """Tìm phần Recommended Songs và nhấn View All cạnh nó"""
        try:
            # In page_source và lưu vào file để debug
            page_source = self.driver.page_source
            debug_file = r"D:\Nam3\hocky2\kiemthu_giuaky\app_music\test\appium_tests\result\page_source.xml"
            os.makedirs(os.path.dirname(debug_file), exist_ok=True)
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(page_source)
            print(f"🔍 Đã lưu cấu trúc giao diện vào: {debug_file}")
            logger.info(f"Đã lưu cấu trúc giao diện vào: {debug_file}")

            # Tìm tiêu đề "Recommended Songs"
            recommended_songs = self.wait.until(
                EC.presence_of_element_located(
                    (AppiumBy.XPATH, "//android.widget.TextView[@text='Recommended Songs']")
                ),
                message="Không tìm thấy tiêu đề Recommended Songs"
            )
            print("✅ Đã tìm thấy phần Recommended Songs.")
            logger.info("Đã tìm thấy phần Recommended Songs.")

            # Nhấn nút "View All" cạnh Recommended Songs
            view_all_button = self.wait.until(
                EC.element_to_be_clickable(
                    (AppiumBy.XPATH, "//android.widget.TextView[@text='View All' and @clickable='true']")
                ),
                message="Không tìm thấy nút View All cạnh Recommended Songs"
            )
            view_all_button.click()
            print("✅ Đã nhấn nút View All cạnh Recommended Songs.")
            logger.info("Đã nhấn nút View All cạnh Recommended Songs.")

            # Đợi danh sách bài hát xuất hiện
            self.wait.until(
                EC.presence_of_element_located((AppiumBy.XPATH, "//android.widget.ListView")),
                message="Không tìm thấy danh sách bài hát sau khi nhấn View All"
            )
            print("✅ Đã tải danh sách bài hát.")
            logger.info("Đã tải danh sách bài hát.")
            return True
        except Exception as e:
            print(f"⚠️ Lỗi khi tìm hoặc nhấn View All cạnh Recommended Songs: {e}")
            logger.warning(f"Lỗi khi tìm hoặc nhấn View All cạnh Recommended Songs: {e}")
            return False

    def select_random_song(self):
        """Chọn một bài hát bất kỳ từ danh sách sau khi nhấn View All"""
        try:
            # Đợi danh sách bài hát xuất hiện trong ListView
            song_elements = self.wait.until(
                EC.presence_of_all_elements_located(
                    (AppiumBy.XPATH, "//android.widget.ListView//android.widget.TextView")
                ),
                message="Không tìm thấy bài hát trong danh sách Recommended Songs"
            )
            if not song_elements:
                raise selenium.common.exceptions.NoSuchElementException("Không tìm thấy bài hát nào trong danh sách")

            # Chọn bài hát đầu tiên trong danh sách
            selected_song = song_elements[0]
            song_name = selected_song.get_attribute("text")
            selected_song.click()
            print(f"✅ Đã chọn bài hát: {song_name}")
            logger.info(f"Đã chọn bài hát: {song_name}")
            return song_name
        except Exception as e:
            print(f"⚠️ Lỗi khi chọn bài hát: {e}")
            logger.error(f"Lỗi khi chọn bài hát: {e}")
            raise

    def play_pause_song(self, song_name):
        """Phát bài hát và tạm dừng sau 5 giây"""
        test_case = "Play-Pause Song"
        try:
            # Đợi giao diện phát nhạc xuất hiện (NowPlayingScreen)
            self.wait.until(
                EC.presence_of_element_located((AppiumBy.XPATH, "//*[contains(@content-desc, 'Now Playing')]")),
                message="Không vào được giao diện phát nhạc"
            )
            print("✅ Đã vào giao diện phát nhạc.")
            logger.info("Đã vào giao diện phát nhạc.")

            # Đợi 5 giây trong khi bài hát đang phát
            print(f"🔊 Đang phát bài hát: {song_name}...")
            logger.info(f"Đang phát bài hát: {song_name}...")
            time.sleep(5)

            # Tìm và nhấn nút Pause
            pause_button = self.wait.until(
                EC.element_to_be_clickable((AppiumBy.XPATH, "//android.widget.Button[@content-desc='Pause']")),
                message="Không tìm thấy nút Pause"
            )
            pause_button.click()
            print("⏸️ Đã tạm dừng bài hát.")
            logger.info("Đã tạm dừng bài hát.")

            # Lưu kết quả thành công vào Excel
            self.save_to_excel(
                test_case=test_case,
                song_name=song_name,
                result="Phát và tạm dừng bài hát thành công",
                status="PASSED"
            )

        except selenium.common.exceptions.TimeoutException as e:
            print(f"⚠️ Timeout khi thực hiện Play-Pause: {e}")
            logger.error(f"Timeout khi thực hiện Play-Pause: {e}")
            self.save_to_excel(
                test_case=test_case,
                song_name=song_name,
                result=f"Timeout: {str(e)}",
                status="FAILED"
            )
            raise
        except selenium.common.exceptions.NoSuchElementException as e:
            print(f"⚠️ Không tìm thấy phần tử khi thực hiện Play-Pause: {e}")
            logger.error(f"Không tìm thấy phần tử khi thực hiện Play-Pause: {e}")
            self.save_to_excel(
                test_case=test_case,
                song_name=song_name,
                result=f"Không tìm thấy phần tử: {str(e)}",
                status="FAILED"
            )
            raise
        except Exception as e:
            print(f"⚠️ Lỗi không xác định khi thực hiện Play-Pause: {e}")
            logger.error(f"Lỗi không xác định khi thực hiện Play-Pause: {e}")
            self.save_to_excel(
                test_case=test_case,
                song_name=song_name,
                result=f"Lỗi: {str(e)}",
                status="FAILED"
            )
            raise

    def test_play_pause_play(self):
        """Tìm Recommended Songs, nhấn View All, chọn một bài hát, phát 5 giây, rồi tạm dừng"""
        self.login()

        # Đảm bảo ứng dụng ở trạng thái foreground
        self.ensure_app_foreground()

        # Tìm phần Recommended Songs và nhấn View All
        if not self.scroll_to_recommended_songs():
            self.fail("❌ Không thể tìm hoặc nhấn View All cạnh Recommended Songs.")

        # Chọn một bài hát bất kỳ
        song_name = self.select_random_song()

        # Thực hiện Play → Pause
        self.play_pause_song(song_name)

    def tearDown(self):
        """Dọn dẹp sau mỗi test case"""
        if self.driver:
            try:
                self.driver.quit()
                print("✅ Đã đóng driver.")
                logger.info("Đã đóng driver.")
            except Exception as e:
                print(f"⚠️ Không thể đóng driver: {e}")
                logger.warning(f"Không thể đóng driver: {e}")

if __name__ == "__main__":
    unittest.main(verbosity=2)